#!/usr/bin/env python3
# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (C) 2026 James Sparenberg
"""
DocuBrowse Excel extractor.

Extracts text and metadata from .xlsx files using openpyxl.

Text extraction strategy:
  - Sheet names collected (used in snippet/subject if no core subject)
  - Each sheet: non-empty cells joined with ' | ' per row, rows as lines
  - read_only=True + data_only=True for speed (no formula evaluation)
  - Stops once TEXT_LIMIT chars collected to avoid giant workbooks

Dependency: pip install openpyxl
"""

from pathlib import Path

_TEXT_LIMIT  = 5000
_MAX_ROWS    = 500   # safety cap per sheet


def extract_xlsx(file_path: str) -> dict:
    result = {
        'title':       None,
        'author':      None,
        'subject':     None,
        'description': '',
        'text':        '',
        'snippet':     '',
        'doc_type':    'xlsx',
        'success':     False,
        'error':       None,
    }
    try:
        import openpyxl

        # read_only for speed; data_only=True returns cached values not formulas
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            # Core properties
            cp = wb.properties
            result['title']   = (getattr(cp, 'title',       '') or '').strip() or None
            result['author']  = (getattr(cp, 'creator',     '') or '').strip() or None
            result['subject'] = (getattr(cp, 'subject',     '') or '').strip() or None
            desc  = (getattr(cp, 'description', '') or '').strip()
            kw    = (getattr(cp, 'keywords',    '') or '').strip()
            result['description'] = desc or kw or ''

            sheet_names = wb.sheetnames

            # Fall back to sheet names as subject if none set
            if not result['subject'] and sheet_names:
                result['subject'] = ', '.join(sheet_names)

            parts = []
            total = 0

            for sheet_name in sheet_names:
                if total >= _TEXT_LIMIT:
                    break
                ws = wb[sheet_name]
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    if total >= _TEXT_LIMIT or row_count >= _MAX_ROWS:
                        break
                    cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
                    if cells:
                        line = ' | '.join(cells)
                        parts.append(line)
                        total += len(line)
                    row_count += 1
        finally:
            wb.close()

        full_text         = '\n'.join(parts)
        result['text']    = full_text[:_TEXT_LIMIT]
        result['snippet'] = full_text[:500]
        result['success'] = True
        return result

    except Exception as exc:
        result['error'] = str(exc)
        return result
